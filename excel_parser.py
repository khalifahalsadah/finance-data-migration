"""Parse the Project Data Completion workbook into normalized Python dicts."""

import openpyxl
from datetime import datetime
from config import (
    SHEET_DATA_START_ROW, PROJECT_FIELDS,
    RESOURCES_PLANNED_COLS, RESOURCES_PLANNED_STATUS_COL,
    RESOURCES_PLANNED_OPP_COL, RESOURCES_PLANNED_QUOT_COL,
    RESOURCES_ACTUAL_COLS, RESOURCES_ACTUAL_STATUS_COL,
    THIRDPARTY_PLANNED_COLS, THIRDPARTY_PLANNED_STATUS_COL,
    THIRDPARTY_PLANNED_OPP_COL, THIRDPARTY_PLANNED_QUOT_COL,
    THIRDPARTY_ACTUAL_COLS, THIRDPARTY_ACTUAL_STATUS_COL,
    DELIVERABLES_COLS, DELIVERABLES_STATUS_COL,
    TOTALS_FIELDS,
)


def _fmt(val):
    """Normalize cell value: dates to string, strip whitespace."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if s.lower() in ('not found', 'none', ''):
        return None
    return s


def _effective(erp_val, corr_val):
    """Return the effective value: corrected if present, else ERP."""
    return _fmt(corr_val) if corr_val is not None and _fmt(corr_val) is not None else _fmt(erp_val)


def _read_pair(ws, row, erp_col, corr_col):
    """Read an ERP/Corrected column pair, return (erp_val, corrected_val, effective_val)."""
    erp = _fmt(ws.cell(row=row, column=erp_col).value)
    corr = _fmt(ws.cell(row=row, column=corr_col).value) if corr_col else None
    eff = corr if corr is not None else erp
    return erp, corr, eff


def parse_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    data = {
        'batches': _parse_batches(wb),
        'project_info': _parse_project_info(wb),
        'totals': _parse_totals(wb),
        'resources_planned': _parse_resources_planned(wb),
        'resources_actual': _parse_resources_actual(wb),
        'thirdparty_planned': _parse_thirdparty_planned(wb),
        'thirdparty_actual': _parse_thirdparty_actual(wb),
        'deliverables': _parse_deliverables(wb),
    }
    return data


def _parse_batches(wb):
    ws = wb['Tracker']
    batches = {}
    for row in range(2, ws.max_row + 1):
        batch_str = ws.cell(row=row, column=1).value
        proj = ws.cell(row=row, column=2).value
        acronym = ws.cell(row=row, column=3).value
        if not proj or not batch_str:
            continue
        try:
            batch_num = int(str(batch_str).replace('Batch ', ''))
        except ValueError:
            continue
        if batch_num not in batches:
            batches[batch_num] = []
        batches[batch_num].append({'id': proj, 'acronym': acronym})
    return batches


def _parse_project_info(wb):
    ws = wb['1. Project Info']
    start_row = SHEET_DATA_START_ROW['1. Project Info']
    result = {}

    for row in range(start_row, ws.max_row + 1):
        proj = ws.cell(row=row, column=1).value
        if not proj:
            continue
        fields = {}
        for erp_col, corr_col, status_col, label, api_field in PROJECT_FIELDS:
            erp_val = _fmt(ws.cell(row=row, column=erp_col).value)
            corr_val = _fmt(ws.cell(row=row, column=corr_col).value)
            status = _fmt(ws.cell(row=row, column=status_col).value)
            fields[api_field] = {
                'label': label,
                'erp_val': erp_val,
                'corrected_val': corr_val,
                'effective_val': corr_val if corr_val is not None else erp_val,
                'status': status,
            }
        result[proj] = fields
    return result


def _parse_totals(wb):
    ws = wb['2. Totals']
    start_row = SHEET_DATA_START_ROW['2. Totals']
    result = {}

    for row in range(start_row, ws.max_row + 1):
        proj = ws.cell(row=row, column=1).value
        if not proj:
            continue
        fields = {}
        for erp_col, corr_col, status_col, label in TOTALS_FIELDS:
            erp_val = _fmt(ws.cell(row=row, column=erp_col).value)
            corr_val = _fmt(ws.cell(row=row, column=corr_col).value)
            status = _fmt(ws.cell(row=row, column=status_col).value)
            fields[label] = {
                'erp_val': erp_val,
                'corrected_val': corr_val,
                'effective_val': corr_val if corr_val is not None else erp_val,
                'status': status,
            }
        result[proj] = fields
    return result


def _parse_detail_sheet(wb, sheet_name, cols_map, status_col, extra_cols=None):
    """Generic parser for detail sheets (3-6) with ERP/Corrected column pairs."""
    ws = wb[sheet_name]
    start_row = SHEET_DATA_START_ROW[sheet_name]
    result = {}

    for row in range(start_row, ws.max_row + 1):
        proj = ws.cell(row=row, column=1).value
        if not proj:
            continue
        if proj not in result:
            result[proj] = []

        row_data = {'_row': row}

        for field, (erp_c, corr_c) in cols_map.items():
            erp_val = _fmt(ws.cell(row=row, column=erp_c).value)
            corr_val = _fmt(ws.cell(row=row, column=corr_c).value) if corr_c else None
            row_data[field] = corr_val if corr_val is not None else erp_val
            row_data[f'{field}_erp'] = erp_val
            row_data[f'{field}_corr'] = corr_val

        row_data['status'] = _fmt(ws.cell(row=row, column=status_col).value)
        row_data['acronym'] = _fmt(ws.cell(row=row, column=2).value)

        if extra_cols:
            for name, col in extra_cols.items():
                row_data[name] = _fmt(ws.cell(row=row, column=col).value)

        result[proj].append(row_data)

    return result


def _parse_resources_planned(wb):
    return _parse_detail_sheet(
        wb, '3. Resources (Planned)',
        RESOURCES_PLANNED_COLS, RESOURCES_PLANNED_STATUS_COL,
        extra_cols={'opportunity': RESOURCES_PLANNED_OPP_COL, 'quotation': RESOURCES_PLANNED_QUOT_COL},
    )


def _parse_resources_actual(wb):
    return _parse_detail_sheet(
        wb, '4. Resources (Actual)',
        RESOURCES_ACTUAL_COLS, RESOURCES_ACTUAL_STATUS_COL,
    )


def _parse_thirdparty_planned(wb):
    return _parse_detail_sheet(
        wb, '5. Third-Party (Planned)',
        THIRDPARTY_PLANNED_COLS, THIRDPARTY_PLANNED_STATUS_COL,
        extra_cols={'opportunity': THIRDPARTY_PLANNED_OPP_COL, 'quotation': THIRDPARTY_PLANNED_QUOT_COL},
    )


def _parse_thirdparty_actual(wb):
    return _parse_detail_sheet(
        wb, '6. Third-Party (Actual)',
        THIRDPARTY_ACTUAL_COLS, THIRDPARTY_ACTUAL_STATUS_COL,
    )


def _parse_deliverables(wb):
    ws = wb['7. Deliverables & Invoices']
    start_row = SHEET_DATA_START_ROW['7. Deliverables & Invoices']
    result = {}

    for row in range(start_row, ws.max_row + 1):
        proj = ws.cell(row=row, column=1).value
        if not proj:
            continue
        if proj not in result:
            result[proj] = []

        cols = DELIVERABLES_COLS
        row_data = {
            '_row': row,
            'deliverable': _fmt(ws.cell(row=row, column=cols['deliverable']).value),
            'amount': _fmt(ws.cell(row=row, column=cols['amount']).value),
            'planned_date': _fmt(ws.cell(row=row, column=cols['planned_date']).value),
            'invoicing_status': _fmt(ws.cell(row=row, column=cols['invoicing_status']).value),
            'actual_date': _fmt(ws.cell(row=row, column=cols['actual_date']).value),
            'invoice_num': _fmt(ws.cell(row=row, column=cols['invoice_num']).value),
            'invoice_date': _fmt(ws.cell(row=row, column=cols['invoice_date']).value),
            'invoice_amount': _fmt(ws.cell(row=row, column=cols['invoice_amount']).value),
            'invoice_status': _fmt(ws.cell(row=row, column=cols['invoice_status']).value),
            'payment_date': _fmt(ws.cell(row=row, column=cols['payment_date']).value),
            'status': _fmt(ws.cell(row=row, column=DELIVERABLES_STATUS_COL).value),
            'acronym': _fmt(ws.cell(row=row, column=2).value),
        }
        result[proj].append(row_data)

    return result
